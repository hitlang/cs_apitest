#-*-coding:utf-8 -*-
#!/usr/bin/python3
# @Author:liulang
import os
import openpyxl
from openpyxl.utils.cell import get_column_letter
import pprint
class ExcelUtil:
    def __init__(self, excel_path , sheetName, *, start_col=1, end_col=None) -> None:
        '''
        :param sheetName: sheet名
        :param start_col: 切片开始列 从1开始
        :param end_col: 切片结束列
        '''
        # wb = openpyxl.load_workbook(os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, "testfiles", "test_cases.xlsx")))
        wb= openpyxl.load_workbook(excel_path)
        self.sheet = wb[sheetName]
        self.start_col_letter = get_column_letter(start_col)
        self.end_col_letter = get_column_letter(end_col)

    def get_dict(self):
        start = "{}{}".format(self.start_col_letter, 1) #包含第一行标题行
        end = "{}{}".format(self.end_col_letter, self.sheet.max_row)
        CELL_SLICE = slice(start, end)
        cells = self.sheet[CELL_SLICE]
        title_cells = [cellobj.value for cellobj in cells[0]]
        ret = []
        for r in cells[1:]:
            m = zip(title_cells, r)
            data_dict = {}
            for k, v in m:
                data_dict[k] = v.value
            ret.append(data_dict)
        return ret

#for test
if __name__ == '__main__':
    excel = ExcelUtil("login", start_col=1, end_col=8)
    r = excel.get_dict()
    pprint.pprint(r)
